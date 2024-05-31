import os
from imbox import Imbox
from datetime import datetime, timedelta
import pandas as pd
from barcodereader import *
from openpyxl import Workbook

username = 'italofarias071@gmail.com'

passwords = open('passwords/token', 'r').read()

host = "imap.gmail.com"
download_folder = "imagens"
#verifica se o diretório imagens existe
os.makedirs(download_folder, exist_ok=True)

mail = Imbox(host, username=username, password=passwords, ssl=True)
# pega emails de 20 dias atras somente os com anexos attachments
messages = mail.messages(date__gt=datetime.today() - timedelta(days=20), raw='has:attachment')

wb = Workbook()

ws = wb.active

r = 1

ws.cell(row=1, column=1).value = "Assunto"
ws.cell(row=1, column=2).value = "Código de Barras"
ws.cell(row=1, column=3).value = "filename"

for (uid, message) in messages:
    if len(message.attachments)>0:
        for attach in message.attachments:
            print(attach)
            att_file = attach["filename"]

            if '.jpg' or '.png' in att_file:
                download_path = f"{download_folder}/{att_file}"

                with open(download_path, 'wb') as fp:
                    fp.write(attach['content'].read())

                try:
                    barcode = BarcodeReader(download_path)
                except:
                    barcode = False

            if not barcode:
                os.remove(download_path)
            else:
                print(message.subject, '-',barcode)
                r += 1
                ws.cell(row=r, column=1).value = message.subject
                ws.cell(row=r, column=2).value = barcode
                ws.cell(row=r, column=3).value = att_file

wb.save("Codigosdebarras.xlsx")
mail.logout()


